"""Claim-bundle parsing and table extraction, vendored verbatim from
claim-bundle-extraction-v2/export_claim_bundle_excels.py.

Left byte-for-byte identical (including its own Excel-export main(), which
the pipeline never calls) so that fixes can be diffed straight across the two
copies. pipeline.py uses BundleContext, build_summary(), the seven
extract_*_rows() functions, AGGREGATION_RULES, aggregate_rows() and
deduplicate_rows().
"""
from __future__ import annotations

import hashlib
import json
import mimetypes
import re
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


# ========================= CONFIGURATION =========================
INPUT_EXCEL = Path("Codes with pipes- to be separated.xlsx")
BUNDLES_ROOT = Path("data/claim-bundle")
OUTPUT_DIR = Path("claim_bundle_review")
N_CLAIMS = 100

# First column in INPUT_EXCEL must contain case_id values such as:
# PMJAY/UK/S/2026/R4/2026070810062219
# registration_id is extracted from the text after the last '/'.
# ================================================================

DOCUMENT_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".gif", ".webp"}


def read_json(path: Path) -> Any | None:
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (OSError, UnicodeDecodeError, json.JSONDecodeError):
        return None


def iter_nodes(value: Any) -> Iterable[Any]:
    yield value
    if isinstance(value, dict):
        for item in value.values():
            yield from iter_nodes(item)
    elif isinstance(value, list):
        for item in value:
            yield from iter_nodes(item)


def norm_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def get_ci(mapping: dict[str, Any], *keys: str) -> Any:
    wanted = {norm_key(key) for key in keys}
    for key, value in mapping.items():
        if norm_key(key) in wanted and value not in (None, "", [], {}):
            return value
    return None


def first_value(obj: Any, *keys: str) -> Any:
    for node in iter_nodes(obj):
        if isinstance(node, dict):
            value = get_ci(node, *keys)
            if value not in (None, "", [], {}):
                return value
    return None


def to_text(value: Any) -> str | None:
    if value in (None, "", [], {}):
        return None
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def to_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = re.sub(r"[^0-9.\-]", "", value)
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None
    return None


def claim_id_from_case_id(value: Any) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip().rstrip("/")
    return text.rsplit("/", 1)[-1] if text else None


def source_context(path: Path, bundle: Path) -> tuple[str | None, str | None, str]:
    rel = path.relative_to(bundle).as_posix()
    parts = rel.split("/")
    side = parts[0] if parts and parts[0] in {"payer", "provider"} else None
    stage = parts[1] if len(parts) > 1 else None
    if stage == "claims":
        stage = "claim"
    return side, stage, rel


CORE_JSON_PATTERNS = [
    ("payer", "claim", "payer/claims/**/*.json"),
    ("payer", "preauthorization", "payer/preauthorization/**/*.json"),
    ("payer", "enhancement", "payer/enhancement/**/*.json"),
    ("provider", "claim", "provider/claim/**/*.json"),
    ("provider", "preauthorization", "provider/preauthorization/**/*.json"),
]


# ============================================================================
# BundleContext
# ----------------------------------------------------------------------------
# The eight extractors used to independently re-glob and re-parse the same JSON
# files (5-7 times per claim) and to call bundle.rglob() once per attachment
# field. BundleContext walks the bundle exactly once, parses each JSON exactly
# once, and exposes reusable indexes so every extractor reads from memory.
# ============================================================================
class BundleContext:
    def __init__(self, bundle: Path):
        self.bundle = bundle
        self._read_cache: dict[Path, Any] = {}
        self._all_files: list[Path] | None = None
        self._json_paths: list[Path] | None = None
        self._files_by_name: dict[str, list[Path]] | None = None
        self._glob_cache: dict[str, list[Path]] = {}
        self._core_jsons: list[tuple[Path, dict[str, Any], str, str]] | None = None
        self._attachment_index: dict[str, list[dict[str, Any]]] | None = None

    def read(self, path: Path) -> Any | None:
        if path not in self._read_cache:
            self._read_cache[path] = read_json(path)
        return self._read_cache[path]

    def all_files(self) -> list[Path]:
        if self._all_files is None:
            self._all_files = [p for p in self.bundle.rglob("*") if p.is_file()]
        return self._all_files

    def files_by_name(self) -> dict[str, list[Path]]:
        if self._files_by_name is None:
            index: dict[str, list[Path]] = {}
            for path in self.all_files():
                index.setdefault(path.name.lower(), []).append(path)
            self._files_by_name = index
        return self._files_by_name

    def json_paths(self) -> list[Path]:
        if self._json_paths is None:
            self._json_paths = [
                p for p in self.all_files()
                if p.suffix.lower() == ".json" and p.name != "manifest.json"
            ]
        return self._json_paths

    def glob(self, pattern: str) -> list[Path]:
        if pattern not in self._glob_cache:
            self._glob_cache[pattern] = sorted(self.bundle.glob(pattern))
        return self._glob_cache[pattern]

    def core_jsons(self) -> list[tuple[Path, dict[str, Any], str, str]]:
        if self._core_jsons is None:
            result: list[tuple[Path, dict[str, Any], str, str]] = []
            seen: set[Path] = set()
            for side, default_stage, pattern in CORE_JSON_PATTERNS:
                for path in self.glob(pattern):
                    if path in seen or path.name == "manifest.json":
                        continue
                    seen.add(path)
                    obj = self.read(path)
                    if not isinstance(obj, dict):
                        continue
                    stage = default_stage
                    if side == "provider" and "enhancement" in path.name.lower():
                        stage = "enhancement"
                    result.append((path, obj, side, stage))
            self._core_jsons = result
        return self._core_jsons

    def resolve_attachment(self, attachment_name: Any, attachment_ref: Any) -> str | None:
        """Resolve an attachment reference to the downloaded EC2 file path."""
        index = self.files_by_name()
        candidates: list[Path] = []

        if attachment_ref:
            ref_name = str(attachment_ref).replace("\\/", "/").rstrip("/").rsplit("/", 1)[-1]
            if ref_name:
                candidates.extend(index.get(ref_name.lower(), []))

        if attachment_name:
            candidates.extend(index.get(str(attachment_name).lower(), []))

        unique = sorted({p.resolve() for p in candidates if p.is_file()})
        return str(unique[0]) if unique else None

    def attachment_index(self) -> dict[str, list[dict[str, Any]]]:
        if self._attachment_index is None:
            index: dict[str, list[dict[str, Any]]] = {}
            for json_path in self.json_paths():
                obj = self.read(json_path)
                if obj is None:
                    continue
                json_path_str = str(json_path.resolve())
                for node in iter_nodes(obj):
                    if not isinstance(node, dict):
                        continue

                    # claimdocuments[] style item: the description (typedesc),
                    # category (type) and MAND code (typeid) live on the item,
                    # while the file lives in a nested "attachments" object.
                    # Link the two so typedesc reaches the document row.
                    nested = node.get("attachments")
                    if isinstance(nested, dict) and ("typedesc" in node or "typeid" in node):
                        nested_ref = get_ci(nested, "attachmentcontent", "attachmentpath")
                        nested_name = get_ci(nested, "attachmentname", "filename")
                        if nested_ref or nested_name:
                            cd_name = str(nested_name or str(nested_ref).rsplit("/", 1)[-1])
                            index.setdefault(cd_name.lower(), []).append({
                                "source_client_s3_path": to_text(nested_ref),
                                "document_type": None,
                                "document_desc": to_text(get_ci(node, "typedesc")),
                                "document_status": to_text(get_ci(node, "status", "documentstatus")),
                                "field_id": None,
                                "form_id": None,
                                "mandatory_code": to_text(get_ci(node, "typeid", "mandatorycode", "mandcode")),
                                "source_json_path": json_path_str,
                            })

                    attachment = get_ci(node, "attachmentcontent", "attachmentpath", "documentreference")
                    name = get_ci(node, "attachmentname", "filename")
                    if not attachment and not name:
                        continue
                    file_name = str(name or str(attachment).rsplit("/", 1)[-1])
                    index.setdefault(file_name.lower(), []).append({
                        "source_client_s3_path": to_text(attachment),
                        "document_type": to_text(get_ci(node, "fieldname", "documenttype", "name", "description")),
                        "document_desc": None,
                        "document_status": to_text(get_ci(node, "documentstatus", "status")),
                        "field_id": to_text(get_ci(node, "fieldid")),
                        "form_id": to_text(get_ci(node, "formid")),
                        "mandatory_code": to_text(get_ci(node, "mandatorycode", "mandcode")),
                        "source_json_path": json_path_str,
                    })
            self._attachment_index = index
        return self._attachment_index


def _ctx(bundle: Path, ctx: BundleContext | None) -> BundleContext:
    return ctx if ctx is not None else BundleContext(bundle)


# Backward-compatible module-level wrappers (delegate to a fresh context).
def iter_core_jsons(bundle: Path) -> Iterable[tuple[Path, dict[str, Any], str, str]]:
    return BundleContext(bundle).core_jsons()


def resolve_local_attachment_path(bundle: Path, attachment_name: Any, attachment_ref: Any) -> str | None:
    return BundleContext(bundle).resolve_attachment(attachment_name, attachment_ref)


def attachment_index(bundle: Path) -> dict[str, list[dict[str, Any]]]:
    return BundleContext(bundle).attachment_index()


def latest_valid_json(ctx: BundleContext, patterns: list[str]) -> tuple[Path | None, Any | None]:
    files: list[Path] = []
    for pattern in patterns:
        files.extend(ctx.glob(pattern))
    files = sorted(set(files), key=lambda p: p.stat().st_mtime if p.exists() else 0)
    for path in reversed(files):
        obj = ctx.read(path)
        if obj is not None:
            return path, obj
    return None, None


def diagnosis_values(obj: Any) -> tuple[str | None, str | None]:
    for node in iter_nodes(obj):
        if not isinstance(node, dict):
            continue
        code = get_ci(node, "diagnosiscode", "icdcode", "code")
        name = get_ci(node, "diagnosisname", "diagnosis", "display", "name")
        keys = {norm_key(k) for k in node}
        if code and name and any("diagnos" in k or "icd" in k for k in keys):
            return to_text(code), to_text(name)
    return (
        to_text(first_value(obj, "diagnosiscode", "icdcode")),
        to_text(first_value(obj, "diagnosisname", "diagnosisdisplay")),
    )


def primary_diagnosis(obj: Any) -> tuple[str | None, str | None]:
    """Pull the primary diagnosis straight from the diagnosis[] array (reliable),
    falling back to the fuzzy node search only if the array is absent.

    The array items use keys code/display/type (no 'diagnosis'/'icd' in the key
    names), which the fuzzy diagnosis_values() heuristic misses.
    """
    items = list_section(obj, "diagnosis") if isinstance(obj, dict) else []
    chosen = None
    for item in items:
        if str(item.get("type", "")).lower() == "primary":
            chosen = item
            break
    if chosen is None and items:
        chosen = items[0]
    if chosen:
        code = to_text(chosen.get("code"))
        name = to_text(chosen.get("display") or chosen.get("name"))
        if code or name:
            return code, name
    return diagnosis_values(obj)


def pick_encounter(*docs: Any) -> dict[str, Any]:
    """Return the first non-empty encounter{} block across the given core docs.

    encounter is populated on every claim/preauth JSON and carries clean patient,
    provider, payer and household identity (unlike the fuzzy first_value scrape).
    """
    for doc in docs:
        if isinstance(doc, dict):
            enc = doc.get("encounter")
            if isinstance(enc, dict) and get_ci(enc, "patientname", "benid", "providerid"):
                return enc
    return {}


def build_summary(case_id: str, claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> dict[str, Any]:
    ctx = _ctx(bundle, ctx)
    _, payer_claim = latest_valid_json(ctx, ["payer/claims/**/*.json"])
    _, payer_preauth = latest_valid_json(ctx, ["payer/preauthorization/**/*.json"])
    _, provider_claim = latest_valid_json(ctx, ["provider/claim/**/*.json"])
    _, provider_preauth = latest_valid_json(ctx, ["provider/preauthorization/**/*.json"])
    _, payer_ben = latest_valid_json(ctx, ["payer/beneficary/**/beneficiary.json", "payer/beneficiary/**/beneficiary.json"])
    _, provider_ben = latest_valid_json(ctx, ["provider/beneficiary/registration/*.json"])

    primary = payer_claim or provider_claim or payer_preauth or provider_preauth or {}
    diagnosis_code, diagnosis_name = primary_diagnosis(primary)
    beneficiary = payer_ben or provider_ben or {}
    encounter = pick_encounter(payer_claim, provider_claim, payer_preauth, provider_preauth)
    amount = primary.get("amount") if isinstance(primary, dict) else None
    amount = amount if isinstance(amount, dict) else {}

    return {
        "registration_id": claim_id,
        "case_id": case_id,
        "provider_id": to_text(get_ci(encounter, "providerid") or first_value(primary, "providerid", "provider_code", "providercode")),
        "provider_name": to_text(get_ci(encounter, "providername")),
        "hospital_code": to_text(first_value(primary, "hospitalcode", "hospitalid", "facilityid") or get_ci(encounter, "providerid")),
        "payer_id": to_text(get_ci(encounter, "payerid")),
        "payer_name": to_text(get_ci(encounter, "payername")),
        "beneficiary_id": to_text(first_value(beneficiary, "beneficiaryid", "benid", "ayushmanid", "uuid") or get_ci(encounter, "benid")),
        "member_id": to_text(first_value(beneficiary, "memberid")),
        "family_id": to_text(first_value(beneficiary, "familyid") or get_ci(encounter, "familyid")),
        "household_id": to_text(get_ci(encounter, "householdid") or first_value(beneficiary, "householdid", "hhid")),
        "patient_name": to_text(get_ci(encounter, "patientname") or first_value(beneficiary, "name")),
        "patient_dob": to_text(get_ci(encounter, "patientdob") or first_value(beneficiary, "dateofbirth", "dob")),
        "patient_gender": to_text(get_ci(encounter, "patientgender") or first_value(beneficiary, "gender")),
        "careplan_id": to_text(get_ci(encounter, "careplanid")),
        "admission_type": to_text(get_ci(primary, "admissiontype")),
        "ipop": to_text(get_ci(primary, "ipop")),
        "diagnosis_code": diagnosis_code,
        "diagnosis_name": diagnosis_name,
        "admission_date": to_text(first_value(primary, "admissiondate", "admissiondatetime")),
        "surgery_date": to_text(first_value(primary, "surgerydate", "treatmentdate")),
        "discharge_date": to_text(first_value(primary, "dischargedate", "dischargedatetime")),
        "requested_amount": to_number(get_ci(amount, "totalpackageamount", "packageamount") or first_value(provider_preauth or payer_preauth, "requestedamount")),
        "approved_amount": to_number(get_ci(amount, "approvedamount") or first_value(payer_preauth or provider_preauth, "approvedamount")),
        "claimed_amount": to_number(get_ci(amount, "claimedamount") or first_value(provider_claim or payer_claim, "claimedamount", "claimamount")),
        "bill_amount": to_number(first_value(provider_claim or payer_claim, "billamount", "hospitalbillamount") or get_ci(amount, "totalamount")),
        "claim_status": to_text(first_value(payer_claim or provider_claim, "claimstatus", "casestatus", "status")),
        "preauth_status": to_text(first_value(payer_preauth or provider_preauth, "preauthstatus", "casestatus", "status")),
    }


def form_instance_from_path(path: Path, bundle: Path) -> str | None:
    parts = path.relative_to(bundle).parts
    if len(parts) >= 4 and parts[0] == "provider" and parts[1] == "forms":
        if path.name == "reports.json":
            return parts[-2]
        return parts[2]
    if len(parts) >= 3 and parts[0] == "payer" and parts[1] == "forms":
        return parts[-2]
    return None


def extract_form_rows(claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> list[dict[str, Any]]:
    ctx = _ctx(bundle, ctx)
    rows: list[dict[str, Any]] = []
    paths = list(ctx.glob("provider/forms/*/*/reports.json"))
    paths += list(ctx.glob("payer/forms/**/*.json"))

    for path in sorted(set(paths)):
        obj = ctx.read(path)
        if not isinstance(obj, dict):
            continue
        side, stage, _ = source_context(path, bundle)
        form_id = to_text(get_ci(obj, "formid"))
        form_name = to_text(get_ci(obj, "formname", "formdesc"))
        form_instance_id = form_instance_from_path(path, bundle)

        groups = get_ci(obj, "groups", "group") or []
        if isinstance(groups, dict):
            groups = [groups]
        if not isinstance(groups, list):
            continue

        for group in groups:
            if not isinstance(group, dict):
                continue
            fields = get_ci(group, "fields", "field") or []
            if isinstance(fields, dict):
                fields = [fields]
            if not isinstance(fields, list):
                continue
            for field in fields:
                if not isinstance(field, dict):
                    continue
                value = get_ci(field, "fieldvalue", "value", "answer")
                attachment_name = get_ci(field, "attachmentname", "filename")
                attachment_ref = get_ci(field, "attachmentcontent", "attachmentpath")
                local_attachment_path = ctx.resolve_attachment(attachment_name, attachment_ref)
                if value in (None, "") and attachment_name:
                    value = attachment_name
                rows.append({
                    "registration_id": claim_id,
                    "source_side": side,
                    "source_stage": stage or "form",
                    "form_id": form_id,
                    "form_name": form_name,
                    "form_instance_id": form_instance_id,
                    "group_id": to_text(get_ci(group, "groupid")),
                    "group_name": to_text(get_ci(group, "groupname")),
                    "field_id": to_text(get_ci(field, "fieldid")),
                    "field_name": to_text(get_ci(field, "fieldname", "label")),
                    "field_value": to_text(value),
                    "field_type": to_text(get_ci(field, "fieldtype", "type")),
                    "is_required": get_ci(field, "isrequired", "required"),
                    "attachment_name": to_text(attachment_name),
                    "attachment_path": local_attachment_path,
                    "source_client_s3_path": to_text(attachment_ref),
                    "source_json_path": str(path.resolve()),
                })
    return rows


def list_section(obj: dict[str, Any], section: str) -> list[dict[str, Any]]:
    value = obj.get(section)
    if value is None:
        return []
    if isinstance(value, dict):
        return [value]
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []


def attachment_values(value: Any) -> tuple[str | None, str | None]:
    if not isinstance(value, dict):
        return None, None
    return (
        to_text(value.get("attachmentname")),
        to_text(value.get("attachmentcontent")),
    )


def extract_diagnosis_rows(claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> list[dict[str, Any]]:
    ctx = _ctx(bundle, ctx)
    rows: list[dict[str, Any]] = []
    for path, obj, side, stage in ctx.core_jsons():
        for sequence, item in enumerate(list_section(obj, "diagnosis"), start=1):
            rows.append({
                "registration_id": claim_id,
                "source_side": side,
                "source_stage": stage,
                "sequence_no": item.get("sno") or sequence,
                "diagnosis_id": to_text(item.get("id")),
                "diagnosis_code": to_text(item.get("code")),
                "diagnosis_name": to_text(item.get("display")),
                "diagnosis_type": to_text(item.get("type")),
                "diagnosis_type_description": to_text(item.get("typedescription")),
                "source_json_path": str(path.resolve()),
            })
    return rows


def extract_care_team_rows(claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> list[dict[str, Any]]:
    ctx = _ctx(bundle, ctx)
    rows: list[dict[str, Any]] = []
    for path, obj, side, stage in ctx.core_jsons():
        for sequence, item in enumerate(list_section(obj, "careteam"), start=1):
            rows.append({
                "registration_id": claim_id,
                "source_side": side,
                "source_stage": stage,
                "sequence_no": item.get("sno") or sequence,
                "doctor_id": to_text(item.get("docid")),
                "doctor_name": to_text(item.get("docname")),
                "doctor_registration_number": to_text(item.get("docregnum")),
                "doctor_qualification": to_text(item.get("docqualification")),
                "doctor_contact_number": to_text(item.get("doccontactnumber")),
                "hpr_id": to_text(item.get("hprid")),
                "source_json_path": str(path.resolve()),
            })
    return rows


def extract_treatment_rows(claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> list[dict[str, Any]]:
    ctx = _ctx(bundle, ctx)
    rows: list[dict[str, Any]] = []
    for path, obj, side, stage in ctx.core_jsons():
        for sequence, item in enumerate(list_section(obj, "treatments"), start=1):
            attachment_name, attachment_path = attachment_values(item.get("attachments"))
            rows.append({
                "registration_id": claim_id,
                "source_side": side,
                "source_stage": stage,
                "sequence_no": item.get("sno") or sequence,
                "item_sequence": item.get("itemsequence"),
                "item_id": to_text(item.get("itemid")),
                "speciality_id": to_text(item.get("specialityid")),
                "treatment_type": to_text(item.get("type")),
                "treatment_type_description": to_text(item.get("typedesc")),
                "procedure_id": to_text(item.get("procedureid")),
                "procedure_code": to_text(item.get("procedurecode")),
                "procedure_name": to_text(item.get("procedurename")),
                "stratification_code": to_text(item.get("procedurestrat")),
                "stratification_name": to_text(item.get("procedurestratname")),
                "procedure_type": to_text(item.get("proceduretype")),
                "program_code": to_text(item.get("programcode")),
                "product_or_service": to_text(item.get("productOrService")),
                "date_on_which": to_text(item.get("dateonwhich")),
                "number_of_days": to_number(item.get("noofdays")),
                "quantity": to_number(item.get("quantity")),
                "approved_quantity": to_number(item.get("approvedquantity")),
                "unit_price": to_number(item.get("unitprice")),
                "base_amount": to_number(item.get("amount")),
                "stratification_amount": to_number(item.get("stratamount")),
                "net_amount": to_number(item.get("netamount")),
                "approved_amount": to_number(item.get("approvedamount")),
                "factor": to_number(item.get("factor")),
                "approved_factor": to_number(item.get("approvedfactor")),
                "length_of_stay": to_number(item.get("los")),
                "is_daycare": to_text(item.get("isdaycare")),
                "status": to_text(item.get("status")),
                "remarks": to_text(item.get("remarks")),
                "reason_code": to_text(item.get("reasoncode")),
                "procedure_category": to_text(item.get("procedurecategoryflag")),
                "attachment_name": attachment_name,
                "attachment_path": ctx.resolve_attachment(attachment_name, attachment_path),
                "source_client_s3_path": attachment_path,
                "source_json_path": str(path.resolve()),
            })
    return rows


def extract_investigation_rows(claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> list[dict[str, Any]]:
    ctx = _ctx(bundle, ctx)
    rows: list[dict[str, Any]] = []
    for path, obj, side, stage in ctx.core_jsons():
        for sequence, item in enumerate(list_section(obj, "investigations"), start=1):
            attachment_name, attachment_path = attachment_values(item.get("attachments"))
            rows.append({
                "registration_id": claim_id,
                "source_side": side,
                "source_stage": stage,
                "source_section": "investigations",
                "sequence_no": item.get("sno") or sequence,
                "item_sequence": item.get("itemsequence"),
                "item_id": to_text(item.get("itemid")),
                "investigation_id": to_text(item.get("id")),
                "investigation_code": to_text(item.get("type")),
                "investigation_name": to_text(item.get("display")),
                "procedure_code": None,
                "procedure_name": None,
                "mandatory_flag": to_text(item.get("mandatoryflag")),
                "quantity": to_number(item.get("quantity")),
                "approved_quantity": to_number(item.get("approvedquantity")),
                "unit_price": to_number(item.get("unitprice")),
                "amount": to_number(item.get("amount")),
                "net_amount": to_number(item.get("netamount")),
                "approved_amount": to_number(item.get("approvedamount")),
                "status": to_text(item.get("status")),
                "remarks": to_text(item.get("remarks")),
                "date_on_which": to_text(item.get("dateonwhich")),
                "attachment_name": attachment_name,
                "attachment_path": ctx.resolve_attachment(attachment_name, attachment_path),
                "source_client_s3_path": attachment_path,
                "document_adjudication": None,
                "source_json_path": str(path.resolve()),
            })

        for proc_sequence, proc in enumerate(list_section(obj, "procedureinvestigationdetails"), start=1):
            details = proc.get("investigationdetails")
            if not isinstance(details, list):
                continue
            for detail_sequence, detail in enumerate(details, start=1):
                if not isinstance(detail, dict):
                    continue
                rows.append({
                    "registration_id": claim_id,
                    "source_side": side,
                    "source_stage": stage,
                        "source_section": "procedureinvestigationdetails",
                    "sequence_no": detail_sequence,
                    "item_sequence": proc_sequence,
                    "item_id": None,
                    "investigation_id": None,
                    "investigation_code": to_text(detail.get("investigationcode")),
                    "investigation_name": to_text(detail.get("investigationname")),
                    "procedure_code": to_text(proc.get("procedurecode")),
                    "procedure_name": to_text(proc.get("procedurename")),
                    "mandatory_flag": None,
                    "quantity": None,
                    "approved_quantity": None,
                    "unit_price": None,
                    "amount": None,
                    "net_amount": None,
                    "approved_amount": None,
                    "status": to_text(proc.get("status")),
                    "remarks": None,
                    "date_on_which": None,
                    "attachment_name": None,
                    "attachment_path": None,
                    "document_adjudication": to_text(detail.get("documentadjudication")),
                    "source_json_path": str(path.resolve()),
                })

        # Implants share the investigation item shape; store them in the same
        # table under source_section = "implants" so all billed non-procedure
        # line items sit together for analysis.
        for sequence, item in enumerate(list_section(obj, "implants"), start=1):
            attachment_name, attachment_path = attachment_values(item.get("attachments"))
            rows.append({
                "registration_id": claim_id,
                "source_side": side,
                "source_stage": stage,
                "source_section": "implants",
                "sequence_no": item.get("sno") or sequence,
                "item_sequence": item.get("itemsequence"),
                "item_id": to_text(item.get("itemid")),
                "investigation_id": to_text(item.get("id")),
                "investigation_code": to_text(item.get("type")),
                "investigation_name": to_text(item.get("display")),
                "procedure_code": None,
                "procedure_name": None,
                "mandatory_flag": to_text(item.get("mandatoryflag")),
                "quantity": to_number(item.get("quantity")),
                "approved_quantity": to_number(item.get("approvedquantity")),
                "unit_price": to_number(item.get("unitprice")),
                "amount": to_number(item.get("amount")),
                "net_amount": to_number(item.get("netamount")),
                "approved_amount": to_number(item.get("approvedamount")),
                "status": to_text(item.get("status")),
                "remarks": to_text(item.get("remarks")),
                "date_on_which": to_text(item.get("dateonwhich")),
                "attachment_name": attachment_name,
                "attachment_path": ctx.resolve_attachment(attachment_name, attachment_path),
                "source_client_s3_path": attachment_path,
                "document_adjudication": None,
                "source_json_path": str(path.resolve()),
            })
    return rows


def extract_package_amount_rows(claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> list[dict[str, Any]]:
    ctx = _ctx(bundle, ctx)
    rows: list[dict[str, Any]] = []
    for path, obj, side, stage in ctx.core_jsons():
        amount = obj.get("amount")
        if not isinstance(amount, dict):
            continue

        calculated = amount.get("calculatedamount")
        if isinstance(calculated, list):
            for sequence, item in enumerate(calculated, start=1):
                if not isinstance(item, dict):
                    continue
                requested = to_number(item.get("requestedamount"))
                approved = to_number(item.get("approvedamount"))
                claimed = to_number(amount.get("claimedamount"))
                rows.append({
                    "registration_id": claim_id,
                    "source_side": side,
                    "source_stage": stage,
                        "sequence_no": item.get("sno") or sequence,
                    "item_sequence": item.get("itemsequence"),
                    "item_id": to_text(item.get("itemid")),
                    "package_code": to_text(item.get("packagecode")),
                    "package_description": to_text(item.get("packagedesc")),
                    "package_type": to_text(item.get("packagetype")),
                    "package_cost": to_number(item.get("packagecost")),
                    "procedure_cost": to_number(item.get("procedurecost")),
                    "stratification_cost": to_number(item.get("stratificationcost")),
                    "quantity": to_number(item.get("quantity")),
                    "approved_quantity": to_number(item.get("approvedquantity")),
                    "percentage_guidelines": to_text(item.get("percentageguidelines")),
                    "hospital_incentive_percentage": to_text(item.get("hospitalincentives")),
                    "incentive_category": to_text(item.get("incentivecategory")),
                    "amount": to_number(item.get("amount")),
                    "net_amount": to_number(item.get("netamount")),
                    "requested_amount": requested,
                    "approved_amount": approved,
                    "claimed_amount": claimed,
                    "status": to_text(item.get("status")),
                    "remarks": to_text(item.get("remarks")),
                    "deductions": to_text(item.get("deductions")),
                    "total_deducted_amount": to_number(item.get("totaldeductedamount")),
                    "reason_code": to_text(item.get("reasoncode")),
                    "source_json_path": str(path.resolve()),
                })

        liabilities = amount.get("liabilities")
        primary_liability = (
            liabilities[0]
            if isinstance(liabilities, list) and liabilities and isinstance(liabilities[0], dict)
            else {}
        )
        rows.append({
            "registration_id": claim_id,
            "source_side": side,
            "source_stage": stage,
            "sequence_no": 0,
            "item_sequence": None,
            "item_id": "CLAIM_TOTAL",
            "package_code": None,
            "package_description": "Claim-level amount totals",
            "package_type": "TOTAL",
            "liability_code": to_text(primary_liability.get("liabilitycode")),
            "wallet_code": to_text(primary_liability.get("walletcode")),
            "wallet_balance_amount": to_number(primary_liability.get("balanceamount")),
            "liability_amount": to_number(primary_liability.get("amount")),
            "liability_approved_amount": to_number(primary_liability.get("approvedamount")),
            "liabilities_json": to_text(liabilities) if liabilities not in (None, "", [], {}) else None,
            "package_cost": to_number(amount.get("packageamount")),
            "procedure_cost": None,
            "stratification_cost": None,
            "quantity": None,
            "approved_quantity": None,
            "percentage_guidelines": None,
            "hospital_incentive_percentage": to_text(amount.get("hospitalincentives")),
            "incentive_category": to_text(amount.get("incentivecategory")),
            "amount": to_number(amount.get("totalamount")),
            "net_amount": to_number(amount.get("netpayable")),
            "requested_amount": to_number(amount.get("totalpackageamount")),
            "approved_amount": to_number(amount.get("approvedamount") or amount.get("amountapproved") or amount.get("claimamountapproved") or amount.get("totalpreauthamountapproved")),
            "claimed_amount": to_number(amount.get("claimedamount")),
            "status": to_text(obj.get("claimstatus") or obj.get("casestatus")),
            "remarks": to_text(obj.get("remarks")),
            "deductions": None,
            "total_deducted_amount": to_number(amount.get("recoveryamount")),
            "reason_code": None,
            "source_json_path": str(path.resolve()),
        })
    return rows


def extract_unmapped_structure_rows(claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> list[dict[str, Any]]:
    """Report unexpected keys inside known medical arrays instead of guessing."""
    ctx = _ctx(bundle, ctx)
    expected = {
        "diagnosis": {"sno", "id", "code", "display", "type", "typedescription"},
        "careteam": {"sno", "docid", "docregnum", "docname", "docqualification", "doccontactnumber", "hprid"},
        "treatments": {"sno", "itemsequence", "itemid", "specialityid", "type", "typedesc", "dateonwhich", "procedureid", "procedurecode", "procedurename", "procedurestrat", "procedurestratname", "stratamount", "proceduretype", "amount", "factor", "approvedfactor", "programcode", "productOrService", "noofdays", "quantity", "approvedquantity", "unitprice", "netamount", "approvedamount", "eligible", "isenhanceable", "los", "ichicode", "isdaycare", "gstapplicable", "gstpercentage", "status", "remarks", "rateflag", "attachments", "reasons", "reasoncode", "deductions", "procedurecategoryflag", "cyclicproc"},
        "investigations": {"sno", "itemsequence", "itemid", "id", "type", "display", "amount", "attachments", "quantity", "mandatoryflag", "approvedquantity", "netamount", "approvedamount", "unitprice", "gstapplicable", "gstpercentage", "status", "remarks", "deductions", "reasons", "reasoncode", "procpolicyinvmpgidpk", "dateonwhich"},
    }
    rows: list[dict[str, Any]] = []
    for path, obj, side, stage in ctx.core_jsons():
        for section, allowed in expected.items():
            for sequence, item in enumerate(list_section(obj, section), start=1):
                unknown = sorted(set(item) - allowed)
                if unknown:
                    rows.append({
                        "registration_id": claim_id,
                        "source_side": side,
                        "source_stage": stage,
                                "section": section,
                        "sequence_no": sequence,
                        "unknown_keys": ", ".join(unknown),
                        "source_json_path": str(path.resolve()),
                    })
    return rows


def sha256(path: Path) -> str | None:
    try:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()
    except OSError:
        return None


def valid_signature(path: Path) -> bool | None:
    try:
        head = path.read_bytes()[:16]
    except OSError:
        return None
    ext = path.suffix.lower()
    if ext == ".pdf":
        return head.startswith(b"%PDF-")
    if ext in {".jpg", ".jpeg"}:
        return head.startswith(b"\xff\xd8\xff")
    if ext == ".png":
        return head.startswith(b"\x89PNG\r\n\x1a\n")
    if ext == ".gif":
        return head.startswith((b"GIF87a", b"GIF89a"))
    if ext == ".webp":
        return head.startswith(b"RIFF") and head[8:12] == b"WEBP"
    return None


def document_type_from_path(path: Path, bundle: Path) -> str | None:
    """Derive a stable document category from the local folder context."""
    rel = path.relative_to(bundle).as_posix().lower()
    if "/investigation_attachment/" in rel:
        return "investigation_attachment"
    if "/attachment/" in rel and "/forms/" in rel:
        return "form_attachment"
    if "/beneficiary/" in rel or "/beneficary/" in rel:
        return "beneficiary_document"
    if "/claim/" in rel or "/claims/" in rel:
        return "claim_document"
    if "/preauthorization/" in rel:
        return "preauthorization_document"
    return "other_document"


def extract_document_rows(claim_id: str, bundle: Path, ctx: BundleContext | None = None) -> list[dict[str, Any]]:
    """Create one row per physical downloaded document.

    JSON references are deduplicated and collapsed into reference_count and
    referenced_by_json_paths so repeated provider snapshots do not create
    duplicate physical-document rows.
    """
    ctx = _ctx(bundle, ctx)
    rows: list[dict[str, Any]] = []
    metadata = ctx.attachment_index()

    for path in sorted(
        p for p in ctx.all_files()
        if p.suffix.lower() in DOCUMENT_EXTENSIONS
    ):
        side, stage, rel = source_context(path, bundle)
        path_mand = re.search(r"(MAND\d+)", rel, re.IGNORECASE)
        raw_matches = metadata.get(path.name.lower(), [])

        unique_matches: dict[tuple[Any, ...], dict[str, Any]] = {}
        for info in raw_matches:
            key = (
                info.get("source_client_s3_path"),
                info.get("document_type"),
                info.get("document_status"),
                info.get("mandatory_code"),
                info.get("source_json_path"),
            )
            unique_matches[key] = info
        matches = list(unique_matches.values())

        json_paths = sorted({m.get("source_json_path") for m in matches if m.get("source_json_path")})
        s3_paths = sorted({m.get("source_client_s3_path") for m in matches if m.get("source_client_s3_path")})
        statuses = sorted({m.get("document_status") for m in matches if m.get("document_status")})
        semantic_types = sorted({m.get("document_type") for m in matches if m.get("document_type")})
        mandatory_codes = sorted({m.get("mandatory_code") for m in matches if m.get("mandatory_code")})
        descriptions = sorted({m.get("document_desc") for m in matches if m.get("document_desc")})

        inferred_type = semantic_types[0] if len(semantic_types) == 1 else document_type_from_path(path, bundle)
        signature_valid = valid_signature(path)

        rows.append({
            "registration_id": claim_id,
            "source_side": side,
            "source_stage": stage,
            "document_type": inferred_type,
            "document_desc": " | ".join(descriptions) if descriptions else None,
            "mandatory_code": mandatory_codes[0] if len(mandatory_codes) == 1 else (path_mand.group(1).upper() if path_mand else None),
            "file_name": path.name,
            "file_extension": path.suffix.lower(),
            "mime_type": mimetypes.guess_type(path.name)[0],
            "source_client_s3_path": " | ".join(s3_paths) if s3_paths else None,
            "source_ec2_path": str(path.resolve()),
            "destination_s3_path": None,
            "file_size_bytes": path.stat().st_size,
            "file_exists_flag": True,
            "valid_file_flag": signature_valid,
            "document_status": " | ".join(statuses) if statuses else None,
            "file_hash_sha256": sha256(path),
            "reference_count": len(matches),
            "referenced_by_json_paths": " | ".join(json_paths) if json_paths else None,
        })
    return rows


def _split_joined(value: Any) -> list[str]:
    """Break a value into its pipe-separated parts.

    A value arriving here can already be pipe-joined - either because the
    source JSON contained one, or because rows were aggregated once before.
    Splitting first means the parts get de-duplicated individually, so a
    second pass can never produce "payer | provider | payer".
    """
    text = str(value).strip()
    if not text:
        return []
    return [part.strip() for part in text.split("|") if part.strip()]


def non_empty(value: Any) -> bool:
    """Whether a cell carries real content. Blank strings, whitespace-only
    strings and empty containers all count as missing."""
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (list, dict, tuple, set)):
        return bool(value)
    return True


def filter_non_empty(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    """Keep only rows where every column in `columns` has a value.

    Used to drop rows that carry no information: a form field with no value,
    a diagnosis with no type description. These are snapshot artefacts - the
    same item is normally also present, populated, in another snapshot - so
    removing them shrinks the table without losing an item.
    """
    if not columns:
        return rows
    return [row for row in rows if all(non_empty(row.get(column)) for column in columns)]


def aggregate_rows(
    rows: list[dict[str, Any]],
    group_cols: list[str],
    join_cols: list[str],
    keep_first_cols: list[str] = (),
    key_cols: list[str] | None = None,
    fallback_key_cols: list[str] | None = None,
) -> list[dict[str, Any]]:
    """Collapse rows to one per distinct identity.

    - `group_cols` are the identity columns kept as-is on the output row.
    - `join_cols` are pipe-joined across the collapsed rows (distinct, sorted) -
      used for provenance like source_side / source_stage / source_client_s3_path,
      and for descriptive columns that vary across snapshots of one item.
    - `keep_first_cols` keep the first row's value (per-item constants that are
      not part of the identity, e.g. file size).
    - `key_cols` overrides which columns decide the grouping, when that differs
      from what is displayed: documents are keyed on their content hash, but
      still show registration_id and the hash on the row.
    - `fallback_key_cols` is used instead of `key_cols` for any row whose key
      columns are not all populated. Without it, every row missing that key
      would collapse into a single bogus group - a document declared in the
      JSON but never downloaded has no hash, and several such documents in one
      claim are different documents, not one.
    - Any column not listed is dropped.

    This removes the payer/provider x stage x snapshot repetition for descriptive
    tables while recording, in one place, which sides/stages carried each value.
    """
    group_cols = list(group_cols)
    join_cols = list(join_cols)
    keep_first_cols = list(keep_first_cols)
    key_cols = list(key_cols) if key_cols else group_cols
    fallback_key_cols = list(fallback_key_cols) if fallback_key_cols else None

    groups: dict[tuple[Any, ...], dict[str, Any]] = {}
    order: list[tuple[Any, ...]] = []
    accumulators: dict[tuple[Any, ...], dict[str, list[str]]] = {}

    for row in rows:
        if all(non_empty(row.get(column)) for column in key_cols):
            key = ("key",) + tuple(row.get(column) for column in key_cols)
        elif fallback_key_cols:
            # Unkeyable row: fall back to the descriptive identity so distinct
            # items stay distinct instead of merging on a shared empty key.
            key = ("fallback",) + tuple(row.get(column) for column in fallback_key_cols)
        else:
            key = ("key",) + tuple(row.get(column) for column in key_cols)

        if key not in groups:
            record = {column: row.get(column) for column in group_cols}
            for column in keep_first_cols:
                record[column] = row.get(column)
            groups[key] = record
            accumulators[key] = {column: [] for column in join_cols}
            order.append(key)
        else:
            record = groups[key]
            # Later rows can fill in an identity/constant column the first row
            # left empty (snapshots are not uniformly populated).
            for column in group_cols + keep_first_cols:
                if not non_empty(record.get(column)) and non_empty(row.get(column)):
                    record[column] = row.get(column)

        acc = accumulators[key]
        for column in join_cols:
            value = row.get(column)
            if not non_empty(value):
                continue
            for part in _split_joined(value):
                if part not in acc[column]:
                    acc[column].append(part)

    result: list[dict[str, Any]] = []
    for key in order:
        record = groups[key]
        for column in join_cols:
            values = accumulators[key][column]
            record[column] = " | ".join(sorted(values)) if values else None
        result.append(record)
    return result


def reduce_by_rule(rows: list[dict[str, Any]], rule: dict[str, Any]) -> list[dict[str, Any]]:
    """Apply one AGGREGATION_RULES entry: drop rows missing required values,
    then collapse what is left. Keeping this here (rather than in the pipeline)
    means the rule's full meaning lives with the rule."""
    rows = filter_non_empty(rows, rule.get("require_non_empty", []))
    return aggregate_rows(
        rows,
        rule["group"],
        rule["join"],
        rule.get("keep_first", []),
        key_cols=rule.get("key"),
        fallback_key_cols=rule.get("fallback_key"),
    )


# Descriptive tables collapsed to one row per logical item, with provenance
# pipe-joined. Keyed by short table name (without the claim_bundle_ prefix).
# Financial tables (treatments/investigations/package_amounts) are NOT here:
# their rows differ by side/status (requested vs approved) and feed the
# side-pivoting analytical views, so they keep the snapshot-dedup behaviour.
AGGREGATION_RULES = {
    # A form field with no value says nothing; the same field is normally
    # present and populated in another snapshot.
    "form_fields": {
        "group": ["registration_id", "form_name", "group_name", "field_name", "field_value", "attachment_name"],
        "join": ["source_side", "source_stage", "source_client_s3_path"],
        "keep_first": [],
        "require_non_empty": ["field_value"],
    },
    # One row per distinct file, identified by content hash: the same physical
    # document is often declared several times under different types, names or
    # descriptions, and those differences belong in the pipe-joined columns
    # rather than in extra rows. Documents whose file was never downloaded have
    # no hash, so they fall back to their descriptive identity.
    "documents": {
        "group": ["registration_id", "file_hash_sha256"],
        "key": ["registration_id", "file_hash_sha256"],
        "fallback_key": ["registration_id", "document_type", "document_desc",
                          "file_name", "file_extension", "mime_type"],
        "join": ["source_side", "source_stage", "document_type", "document_desc",
                  "file_name", "file_extension", "mime_type", "mandatory_code",
                  "source_client_s3_path", "document_status"],
        "keep_first": ["file_size_bytes", "file_exists_flag", "valid_file_flag", "reference_count"],
    },
    # Diagnosis rows without a type description are snapshot noise - the same
    # diagnosis appears elsewhere with its description.
    "diagnosis": {
        "group": ["registration_id", "diagnosis_code", "diagnosis_name", "diagnosis_type", "diagnosis_id", "diagnosis_type_description"],
        "join": ["source_side", "source_stage"],
        "keep_first": [],
        "require_non_empty": ["diagnosis_type_description"],
    },
    # One row per doctor on the claim. The registration number is the doctor's
    # real identity; ids, qualifications and contact numbers vary between
    # snapshots of the same person, so they are pipe-joined rather than
    # splitting one doctor across several rows.
    "care_team": {
        "group": ["registration_id", "doctor_name", "doctor_registration_number"],
        "join": ["source_side", "source_stage", "doctor_id", "doctor_qualification",
                  "doctor_contact_number", "hpr_id"],
        "keep_first": [],
    },
}


def deduplicate_rows(rows: list[dict[str, Any]], ignore_columns: set[str]) -> list[dict[str, Any]]:
    """Drop repeated snapshot rows while preserving source-side/stage differences."""
    seen: set[tuple[tuple[str, str], ...]] = set()
    result: list[dict[str, Any]] = []
    for row in rows:
        key = tuple(sorted(
            (column, json.dumps(value, sort_keys=True, ensure_ascii=False, default=str))
            for column, value in row.items()
            if column not in ignore_columns
        ))
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def save_excel(rows: list[dict[str, Any]], file_name: str, sheet_name: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / file_name
    df = pd.DataFrame(rows)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column in ws.columns:
            letter = column[0].column_letter
            max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column[:300])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    print(f"Created {path} ({len(df)} rows)")


def selected_claims() -> list[tuple[str, str, Path]]:
    df = pd.read_excel(INPUT_EXCEL, usecols=[0], dtype=str)
    selected: list[tuple[str, str, Path]] = []
    seen: set[str] = set()

    for raw_case_id in df.iloc[:, 0].dropna():
        claim_id = claim_id_from_case_id(raw_case_id)
        if not claim_id or claim_id in seen:
            continue
        seen.add(claim_id)

        bundle = BUNDLES_ROOT / claim_id
        if not bundle.is_dir():
            continue

        manifest = read_json(bundle / "manifest.json") or {}
        if manifest.get("status") not in {None, "completed", "completed_with_errors"}:
            continue

        usable_jsons = [p for p in bundle.rglob("*.json") if p.name != "manifest.json" and read_json(p) is not None]
        if not usable_jsons:
            continue

        selected.append((str(raw_case_id), claim_id, bundle))
        if len(selected) >= N_CLAIMS:
            break

    if len(selected) < N_CLAIMS:
        raise RuntimeError(
            f"Only {len(selected)} usable claim bundles found; requested {N_CLAIMS}."
        )
    return selected


def main() -> None:
    claims = selected_claims()

    tables: dict[str, list[dict[str, Any]]] = {
        "summary": [],
        "forms": [],
        "documents": [],
        "diagnosis": [],
        "treatments": [],
        "investigations": [],
        "care_team": [],
        "package_amounts": [],
        "unmapped_structures": [],
    }

    for index, (case_id, claim_id, bundle) in enumerate(claims, start=1):
        print(f"[{index}/{len(claims)}] {claim_id}")
        ctx = BundleContext(bundle)
        tables["summary"].append(build_summary(case_id, claim_id, bundle, ctx))
        tables["forms"].extend(extract_form_rows(claim_id, bundle, ctx))
        tables["documents"].extend(extract_document_rows(claim_id, bundle, ctx))
        tables["diagnosis"].extend(extract_diagnosis_rows(claim_id, bundle, ctx))
        tables["treatments"].extend(extract_treatment_rows(claim_id, bundle, ctx))
        tables["investigations"].extend(extract_investigation_rows(claim_id, bundle, ctx))
        tables["care_team"].extend(extract_care_team_rows(claim_id, bundle, ctx))
        tables["package_amounts"].extend(extract_package_amount_rows(claim_id, bundle, ctx))
        tables["unmapped_structures"].extend(extract_unmapped_structure_rows(claim_id, bundle, ctx))

    for short, rule in AGGREGATION_RULES.items():
        tables[short] = aggregate_rows(tables[short], rule["group"], rule["join"], rule["keep_first"])
    tables["treatments"] = deduplicate_rows(tables["treatments"], {"source_json_path", "sequence_no"})
    tables["investigations"] = deduplicate_rows(tables["investigations"], {"source_json_path", "sequence_no", "item_sequence"})
    tables["package_amounts"] = deduplicate_rows(tables["package_amounts"], {"source_json_path", "sequence_no", "item_sequence"})

    save_excel(tables["summary"], "claim_bundle_summary.xlsx", "claim_bundle_summary")
    save_excel(tables["forms"], "claim_bundle_form_fields.xlsx", "claim_bundle_form_fields")
    save_excel(tables["documents"], "claim_bundle_documents.xlsx", "claim_bundle_documents")
    save_excel(tables["diagnosis"], "claim_bundle_diagnosis.xlsx", "claim_bundle_diagnosis")
    save_excel(tables["treatments"], "claim_bundle_treatments.xlsx", "claim_bundle_treatments")
    save_excel(tables["investigations"], "claim_bundle_investigations.xlsx", "claim_bundle_investigations")
    save_excel(tables["care_team"], "claim_bundle_care_team.xlsx", "claim_bundle_care_team")
    save_excel(tables["package_amounts"], "claim_bundle_package_amounts.xlsx", "claim_bundle_package_amounts")
    save_excel(tables["unmapped_structures"], "claim_bundle_unmapped_structures.xlsx", "unmapped_structures")


if __name__ == "__main__":
    main()
